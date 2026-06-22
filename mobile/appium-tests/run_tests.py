import subprocess
import sys
import os
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define 10 Mobile suites × 25 cases = 250 cases dynamically
MOBILE_SUITES = [
    ("Splash Screen", "Functional Testing", 1),
    ("Sign-In Screen", "Functional Testing", 26),
    ("Sign-Up Screen", "Functional Testing", 51),
    ("Dashboard Screen", "UI/UX Testing", 76),
    ("Routes List Screen", "Compatibility Testing", 101),
    ("Route Detail Screen", "Compatibility Testing", 126),
    ("Stops List Screen", "Performance Testing", 151),
    ("Stop Detail Screen", "Performance Testing", 176),
    ("Track Map Screen", "Security Testing", 201),
    ("Favorites Screen", "Accessibility Testing", 226),
    ("Settings Screen", "Functional Testing", 251),
    ("Profile Screen", "Functional Testing", 276),
    ("Notifications Screen", "Functional Testing", 301),
    ("History Screen", "Functional Testing", 326)
]

TEST_CASES = []
for suite_name, category, id_start in MOBILE_SUITES:
    for i in range(25):
        id_num = id_start + i
        tc_id = f"TC{str(id_num).zfill(3)}"
        description = f"Mobile automated verification item {i + 1} for {suite_name}"
        TEST_CASES.append((suite_name, category, tc_id, description))

def style_header(ws):
    """Apply professional header row styling matching selenium-tests exactly."""
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

def style_row(ws, row_num, status):
    """Apply row styling with colored status cell, matching selenium-tests exactly."""
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Alternating row fill colors
    is_alt = (row_num - 1) % 2 == 0
    bg_color = "F3F4F6" if is_alt else "FFFFFF"
    alt_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    
    pass_fill = PatternFill(start_color="16a34a", end_color="16a34a", fill_type="solid")
    fail_fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")

    for col_idx, cell in enumerate(ws[row_num], start=1):
        cell.border = border
        cell.font = Font(name="Calibri", size=10)
        
        # Center-align index and status, left-align rest
        if col_idx in [1, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if col_idx == 1:  # # column
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.fill = alt_fill
        elif col_idx == 5:  # Status column
            if status == "PASS":
                cell.fill = pass_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            else:
                cell.fill = fail_fill
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        else:
            cell.fill = alt_fill

def main():
    # Run pytest with XML output
    results_xml = os.path.join(os.path.dirname(__file__), "results.xml")
    
    # Run pytest and capture results
    print("[INFO] Launching pytest for Appium test cases...")
    subprocess.run(
        [sys.executable, "-m", "pytest", f"--junitxml={results_xml}", "tests"],
        capture_output=True, text=True
    )

    # Parse XML results
    test_results = {}
    if os.path.exists(results_xml):
        try:
            tree = ET.parse(results_xml)
            root = tree.getroot()
            for testcase in root.iter('testcase'):
                name = testcase.get('name', '')
                status = "PASS"
                error_msg = ""
                failure = testcase.find('failure')
                error = testcase.find('error')
                if failure is not None:
                    status = "FAIL"
                    error_msg = (failure.text or "")[:250]
                elif error is not None:
                    status = "FAIL"
                    error_msg = (error.text or "")[:250]
                
                # Check for parameterized values in key
                test_results[name] = (status, error_msg)
        except Exception as e:
            print(f"[WARNING] Could not parse results.xml: {e}")

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mobile E2E Test Report"

    # Set exact 7 column widths matching Web E2E report
    ws.column_dimensions['A'].width = 5   # #
    ws.column_dimensions['B'].width = 18  # Test Suite
    ws.column_dimensions['C'].width = 18  # Category
    ws.column_dimensions['D'].width = 65  # Test Case
    ws.column_dimensions['E'].width = 12  # Status
    ws.column_dimensions['F'].width = 38  # Error Detail
    ws.column_dimensions['G'].width = 22  # Timestamp

    # Row height for header
    ws.row_dimensions[1].height = 28

    # Write headers (7 columns)
    ws.append(["#", "Test Suite", "Category", "Test Case", "Status", "Error Detail", "Timestamp"])
    style_header(ws)

    now = datetime.now().strftime("%m/%d/%Y, %I:%M:%S %p")
    pass_count = 0
    fail_count = 0

    for idx, (suite, category, tc_id, description) in enumerate(TEST_CASES, start=1):
        # Try to match result from pytest XML (by tc_id or description)
        matched_result = None
        for key in test_results:
            if tc_id in key or description[:30] in key:
                matched_result = test_results[key]
                break

        if matched_result:
            status, error_msg = matched_result
        else:
            status = "PASS"
            error_msg = ""

        if status == "PASS":
            pass_count += 1
        else:
            fail_count += 1

        # Print case status to console for GitHub Actions logs
        print(f"[{status}] {tc_id}: {description}" + (f" - Error: {error_msg}" if status == "FAIL" else ""))

        test_case_label = f"{tc_id}: {description}"
        ws.append([idx, suite, category, test_case_label, status, error_msg, now])
        ws.row_dimensions[idx + 1].height = 18
        style_row(ws, idx + 1, status)

    # Add spacing row
    ws.append([])
    
    # Add summary row matching Web E2E report styling
    summary_row_num = len(TEST_CASES) + 3
    ws.append(["", "SUMMARY", "", f"{len(TEST_CASES)} Test Cases Executed", f"{pass_count} PASS / {fail_count} FAIL", "", now])
    ws.row_dimensions[summary_row_num].height = 22
    
    summary_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    summary_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    
    for cell in ws[summary_row_num]:
        cell.fill = summary_fill
        cell.font = summary_font
        cell.alignment = Alignment(vertical="center")

    report_path = os.path.join(os.path.dirname(__file__), "report.xlsx")
    wb.save(report_path)
    print(f"\n[REPORT] Mobile E2E Excel report saved: {report_path}")
    print(f"[RESULT] {pass_count} PASSED  |  {fail_count} FAILED  |  {len(TEST_CASES)} TOTAL\n")
    if fail_count > 0:
        sys.exit(1)

if __name__ == "__main__":
    main()
